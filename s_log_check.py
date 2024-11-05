import paramiko
import openpyxl
from datetime import datetime, timedelta
import os
import re
import logging


# 로그 출력을 위한 함수
def log_message(message):
    print(f"[LOG] {message}")

log_message("로그 검수 확인 중")
logging.info("로그 검수 확인 중")

# SSH 서버 정보
hostname = '192.168.2.214'
port = 7795
username = 'root'
password = 'dbsafer00'
log_file_path = '/usr/local/apache/logs/saferuas/saferuas.log'

# 엑셀 파일 경로
excel_path = "C:/Users/pnpadmin/Desktop/SAFERUAS TCL_v2.5_20240919.xlsx"

# 오늘 날짜로 시트 이름 생성
current_datetime = datetime.now()
sheet_name = current_datetime.strftime("%Y%m%d%H%M%S")  # 예: 20241029142124

# SSH 클라이언트 설정 및 서버 접속
ssh_client = paramiko.SSHClient()
ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
ssh_client.connect(hostname, port=port, username=username, password=password)

# 로그 파일의 최신 10줄을 추출하는 명령어
command = f"tail -n 10 {log_file_path}"
stdin, stdout, stderr = ssh_client.exec_command(command)

# 명령어 실행 결과를 가져옴
saferuas_log_lines = stdout.readlines()

# SSH 연결 종료
ssh_client.close()

# 엑셀 파일 열기 또는 새로 생성
if os.path.exists(excel_path):
    workbook = openpyxl.load_workbook(excel_path)
else:
    workbook = openpyxl.Workbook()

# 새로운 시트 생성
worksheet = workbook.create_sheet(title=sheet_name)

# B2 셀에 "saferuas 로그" 작성
worksheet["B2"] = "saferuas 로그"

# B3 셀부터 saferuas 로그 한 줄씩 작성
for row_idx, line in enumerate(saferuas_log_lines, start=3):  # B3 부터 시작
    worksheet.cell(row=row_idx, column=2, value=line.strip())

# "PCAsisst 로그"를 위해 두 칸 아래로 이동
worksheet.cell(row=row_idx + 2, column=2, value="PCAsisst 로그")
pca_log_start_row = row_idx + 3

# PCAsisst 로그 파일 경로 설정 (오늘 날짜 기준 파일만 사용)
pca_log_dir = "C:/Users/pnpadmin/AppData/Roaming/DSASSIST/log"
today_log_file = f"{current_datetime.strftime('%Y-%m-%d')}.log"
log_file_path = os.path.join(pca_log_dir, today_log_file)

# 현재 시간 기준 7초 이내 로그 추출
time_threshold = current_datetime - timedelta(seconds=7)

# 오늘 날짜의 로그 파일이 존재하는지 확인 후 처리
if os.path.exists(log_file_path):
    with open(log_file_path, "r", encoding="cp949") as file:  # 인코딩을 cp949로 변경
        for line in file:
            # 로그 시간 추출 정규식
            match = re.match(r"\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\.\d{3}\]", line)
            if match:
                log_time_str = match.group(1)
                log_time = datetime.strptime(log_time_str, "%Y-%m-%d %H:%M:%S")
                # 7초 이내 로그만 기록
                if log_time >= time_threshold:
                    worksheet.cell(row=pca_log_start_row, column=2, value=line.strip())
                    pca_log_start_row += 1


# 엑셀 파일 저장
workbook.save(excel_path)

print(f"{excel_path}에 {sheet_name} 시트가 생성되고 로그가 작성되었습니다.")
logging.info(f"{excel_path}에 {sheet_name} 시트가 생성되고 로그가 작성되었습니다.")

    # 현재 시간 기준으로 최근 5초 이내에 생성된 로그에서 "사용자 로그인 완료" 확인
try:
    log_message("PCAsisst 로그 확인중")
    logging.info("PCAsisst 로그 확인중")
    log_dir = r"C:\Users\pnpadmin\AppData\Roaming\DSASSIST\log"
    log_found = False
    now = datetime.now()
    five_seconds_ago = now - timedelta(seconds=5)

    # 최신 로그 파일을 찾기
    log_files = sorted(os.listdir(log_dir), reverse=True)
    for log_file in log_files:
        log_path = os.path.join(log_dir, log_file)
        with open(log_path, 'r', encoding='cp949') as file:  # 'cp949' 인코딩으로 파일을 읽음
            lines = file.readlines()
            # 로그 파일의 각 줄을 확인
            for line in lines[::-1]:  # 가장 최신 로그부터 탐색
                try:
                    # 로그에 포함된 시간을 추출하여 최근 5초 이내인지 확인
                    timestamp_str = line.split("]")[0].strip("[")
                    log_time = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S.%f")
                    if log_time < five_seconds_ago:
                        break
                    # "사용자 로그인 완료"가 포함된 로그가 있는지 확인
                    if "사용자 로그인 완료" in line:
                        log_message("PCAsisst 로그 검수 확인 완료")
                        logging.info("PCAsisst 로그 검수 확인 완료")
                        log_found = True
                        break
                except ValueError:
                    continue
            if log_found:
                break

except Exception as e:
    log_message(f"PCAsisst 로그 확인 실패 : {e}")
    logging.info(f"PCAsisst 로그 확인 실패 : {e}")
    exit()